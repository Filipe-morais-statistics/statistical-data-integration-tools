try:
    input_file
except NameError:
    print("Erro na obtenção do input file. Contactar administrador do sistema.")
    import sys

    # Para testes
    sys.path.append(
        r"G:\aaa\3. Python\tabulizer"
    )
    input_file = r"G:\aaa\1. Ficheiros originais\Carteira Títulos source - Junho 2025.xlsx"
    data = {}

# Importação do package
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from tabulizer import *
import tabulizer

# Guarda a data da informação
wb = load_workbook(input_file)
ws = wb.active
ano = ws["H10"].value.year
mes = ws["H10"].value.month

# Lê o ficheiro
df = pd.read_excel(input_file, header=11)

# Apaga colunas
df.drop(df.columns[[0, 2, 3, 4, 5, 6, 8]], axis=1, inplace=True)

# Altera o nome das colunas
df.columns = ["Designacao", "Quantidade/Valor Nominal", "Valor de Mercado"]

# Elimina registos dos totais e dos NA
df.drop(df[df.Designacao == "SUB TOTAL"].index, inplace=True)
df.drop(df[df.Designacao == "TOTAL"].index, inplace=True)
df = df.dropna(thresh=2, subset=["Quantidade/Valor Nominal", "Valor de Mercado"])

# Agrega valores na mesma coluna e cria coluna tipo de valor
df = pd.melt(
    df,
    value_vars=["Quantidade/Valor Nominal", "Valor de Mercado"],
    id_vars=["Designacao"],
)

# Altera o nome das colunas
df.columns = ["Designacao", "Tipo_Metrica", "Valor"]

# Adiciona coluna com periodo
df.insert(1, "Ano", ano)
df.insert(2, "Mes", mes)
df["Mes2"] = np.where(df["Mes"] < 10, "0", "")
df["Periodo"] = df["Ano"].astype(str) + "-" + df["Mes2"] + df["Mes"].astype(str)

# altera ordem das colunas
col_names_2 = ["Periodo", "Designacao", "Tipo_Metrica", "Valor"]
df = df.reindex(columns=col_names_2)

# altera ordem das linhas
df = df.sort_values(["Designacao", "Tipo_Metrica"], ascending=[False, False])

# Apenas para testes
# df.to_excel(r"G:\aaa\3. Python\tabela.xlsx")

# Para verificar quais as colunas do dataframe, que devem coincidir com os campos definidos na metainformação
print("Foi criado uma estrutura de dados com os seguintes campos: ", df.columns)

# output
data["CGA_CarteiraTitulos"] = df

#df.to_excel (r"G:\aaa\3. Python\Output_source.xlsx")
